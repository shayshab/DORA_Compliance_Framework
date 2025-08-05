#!/usr/bin/env python3
"""
DORA Compliance Framework Excel Generator
Converts CSV files into a comprehensive Excel workbook for DORA compliance tracking
"""

import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.utils import get_column_letter
import datetime

def create_dora_excel_workbook():
    """Create comprehensive DORA Compliance Framework Excel workbook"""
    
    # Define file paths
    base_path = "/Users/shayshabazad/Files/project/ai project/ai generated/Dora"
    csv_files = {
        'DORA Compliance': 'DORA_Compliance_Framework.csv',
        'Risk Scoring': 'Risk_Scoring_Matrix.csv',
        'Implementation': 'Implementation_Roadmap.csv',
        'KRI & KPI': 'KRI_KPI_Dashboard.csv',
        'Third Party Risk': 'Third_Party_Risk_Register.csv',
        'Incident Management': 'Incident_Management_Log.csv'
    }
    
    # Create new workbook
    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Define colors for styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    
    # Status colors
    status_colors = {
        'Not Started': PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        'Planning': PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid"),
        'In Progress': PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid"),
        'Completed': PatternFill(start_color="95E1D3", end_color="95E1D3", fill_type="solid"),
        'Critical': PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        'High': PatternFill(start_color="FFB347", end_color="FFB347", fill_type="solid"),
        'Medium': PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid"),
        'Low': PatternFill(start_color="95E1D3", end_color="95E1D3", fill_type="solid")
    }
    
    # Process each CSV file
    for sheet_name, csv_file in csv_files.items():
        csv_path = os.path.join(base_path, csv_file)
        
        if os.path.exists(csv_path):
            # Read CSV file
            df = pd.read_csv(csv_path)
            
            # Create worksheet
            ws = wb.create_sheet(title=sheet_name)
            
            # Add data to worksheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Style the header row
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set column width with limits
                adjusted_width = min(max_length + 2, 50)
                adjusted_width = max(adjusted_width, 10)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Apply conditional formatting based on sheet content
            if sheet_name == 'DORA Compliance':
                # Color code implementation status
                status_col = None
                priority_col = None
                
                for idx, cell in enumerate(ws[1], 1):
                    if cell.value == 'Implementation Status':
                        status_col = idx
                    elif cell.value == 'Priority':
                        priority_col = idx
                
                if status_col:
                    status_column_letter = get_column_letter(status_col)
                    for row in range(2, ws.max_row + 1):
                        cell = ws[f'{status_column_letter}{row}']
                        if cell.value in status_colors:
                            cell.fill = status_colors[cell.value]
                
                if priority_col:
                    priority_column_letter = get_column_letter(priority_col)
                    for row in range(2, ws.max_row + 1):
                        cell = ws[f'{priority_column_letter}{row}']
                        if cell.value in status_colors:
                            cell.fill = status_colors[cell.value]
            
            # Freeze panes (first row and first column)
            ws.freeze_panes = 'B2'
            
            # Add filters to the data
            ws.auto_filter.ref = ws.dimensions
    
    # Create Executive Summary worksheet
    exec_summary = wb.create_sheet(title="Executive Summary", index=0)
    
    # Executive Summary content
    summary_data = [
        ["DORA Compliance Framework - Executive Summary"],
        [""],
        ["Report Date:", datetime.datetime.now().strftime("%B %d, %Y")],
        [""],
        ["Overall Compliance Status"],
        ["Total Requirements:", "75"],
        ["Not Started:", "65 (87%)"],
        ["In Progress:", "8 (11%)"],
        ["Completed:", "2 (3%)"],
        [""],
        ["Risk Assessment"],
        ["Critical Priority Items:", "15"],
        ["High Priority Items:", "35"],
        ["Medium Priority Items:", "20"],
        ["Low Priority Items:", "5"],
        [""],
        ["Key Implementation Dates"],
        ["DORA Effective Date:", "January 17, 2025"],
        ["Days Remaining:", "165"],
        ["Next Major Milestone:", "Board Approval - October 15, 2024"],
        [""],
        ["Critical Actions Required"],
        ["1. Immediate board engagement and approval"],
        ["2. Establish DORA program management office"],
        ["3. Complete ICT asset inventory"],
        ["4. Implement incident response procedures"],
        ["5. Enhance third-party risk management"],
        [""],
        ["Budget Requirements (Estimated)"],
        ["Technology Infrastructure:", "€2,500,000"],
        ["External Consulting:", "€1,200,000"],
        ["Staff Training:", "€300,000"],
        ["Regulatory Compliance:", "€500,000"],
        ["Total Estimated Cost:", "€4,500,000"],
        [""],
        ["Resource Requirements"],
        ["Full-time Program Manager:", "1"],
        ["IT Risk Specialists:", "3"],
        ["Security Analysts:", "2"],
        ["Compliance Officers:", "2"],
        ["Business Analysts:", "2"],
        [""],
        ["Key Risks"],
        ["- Insufficient time for full implementation"],
        ["- Resource constraints and competing priorities"],
        ["- Third-party vendor readiness"],
        ["- Regulatory interpretation uncertainty"],
        ["- Integration with existing systems"]
    ]
    
    # Add summary data
    for row_data in summary_data:
        exec_summary.append(row_data)
    
    # Style executive summary
    exec_summary['A1'].font = Font(size=16, bold=True, color="366092")
    exec_summary['A1'].alignment = Alignment(horizontal="center")
    exec_summary.merge_cells('A1:D1')
    
    # Style section headers
    section_headers = [5, 11, 17, 23, 29, 36, 42]
    for row_num in section_headers:
        cell = exec_summary[f'A{row_num}']
        cell.font = Font(bold=True, size=12, color="366092")
        cell.fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
    
    # Auto-adjust column widths for executive summary
    for column in exec_summary.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 40)
        adjusted_width = max(adjusted_width, 15)
        exec_summary.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    output_file = os.path.join(base_path, "DORA_Compliance_Framework.xlsx")
    wb.save(output_file)
    print(f"DORA Compliance Framework Excel file created: {output_file}")
    return output_file

if __name__ == "__main__":
    create_dora_excel_workbook()
